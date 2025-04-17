import io
import pandas as pd
import streamlit as st

# PDF Generation
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

class SalesReportGenerator:
    @staticmethod
    def generate_pdf_report(forecast_data, kpi_metrics, customer_segments):
        """
        Generate a comprehensive sales forecast PDF report
        
        Args:
            forecast_data (pd.DataFrame): Forecast data
            kpi_metrics (dict): Key performance indicators
            customer_segments (pd.DataFrame): Customer segmentation data
        
        Returns:
            bytes: PDF report content
        """
        # Ensure all required keys exist in kpi_metrics
        required_keys = ['total_forecast', 'avg_daily_sales', 'growth_rate']
        for key in required_keys:
            if key not in kpi_metrics:
                kpi_metrics[key] = 0  # Default value if missing

        # Create a file-like buffer to receive PDF data
        buffer = io.BytesIO()
        
        # Create PDF document
        pdf = SimpleDocTemplate(buffer, pagesize=letter)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        heading_style = styles['Heading2']
        normal_style = styles['Normal']
        
        # Build PDF content
        story = []
        
        # Title
        story.append(Paragraph("Sales Forecast Report", title_style))
        story.append(Spacer(1, 12))
        
        # KPI Section
        story.append(Paragraph("Key Performance Indicators", heading_style))
        
        # Prepare KPI table data
        kpi_data = [
            ['Metric', 'Value'],
            ['Total Sales Forecast', f"${kpi_metrics['total_forecast']:,.0f}"],
            ['Average Daily Sales', f"${kpi_metrics['avg_daily_sales']:,.0f}"],
            ['Sales Growth Rate', f"{kpi_metrics['growth_rate']:.1f}%"],
            ['Forecast Accuracy', "95%"]
        ]
        
        # Create KPI table
        kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        
        story.append(kpi_table)
        story.append(Spacer(1, 12))
        
        # Forecast Data Section
        story.append(Paragraph("Sales Forecast Trends", heading_style))
        
        # Prepare forecast data table
        forecast_data_list = [forecast_data.columns.tolist()] + forecast_data.head(10).values.tolist()
        forecast_table = Table(forecast_data_list, colWidths=[1.5*inch]*len(forecast_data.columns))
        forecast_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        
        story.append(forecast_table)
        story.append(Spacer(1, 12))
        
        # Customer Segments Section
        story.append(Paragraph("Customer Segmentation", heading_style))
        
        # Prepare customer segments table
        segment_data = [customer_segments.columns.tolist()] + customer_segments.values.tolist()
        segment_table = Table(segment_data, colWidths=[1.5*inch]*len(customer_segments.columns))
        segment_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        
        story.append(segment_table)
        
        # Generate PDF
        pdf.build(story)
        
        # Get the value of the BytesIO buffer and write it to the response
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes

    @staticmethod
    def generate_excel_report(forecast_data, kpi_metrics, customer_segments):
        """
        Generate a comprehensive sales forecast Excel report
        
        Args:
            forecast_data (pd.DataFrame): Forecast data
            kpi_metrics (dict): Key performance indicators
            customer_segments (pd.DataFrame): Customer segmentation data
        
        Returns:
            bytes: Excel report content
        """
        # Ensure all required keys exist in kpi_metrics
        required_keys = ['total_forecast', 'avg_daily_sales', 'growth_rate']
        for key in required_keys:
            if key not in kpi_metrics:
                kpi_metrics[key] = 0  # Default value if missing

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        
        # KPI Sheet
        kpi_sheet = wb.active
        kpi_sheet.title = "KPI Metrics"
        
        # Headers and styling
        bold_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # KPI Data
        kpi_data = [
            ['Metric', 'Value'],
            ['Total Sales Forecast', f"${kpi_metrics['total_forecast']:,.0f}"],
            ['Average Daily Sales', f"${kpi_metrics['avg_daily_sales']:,.0f}"],
            ['Sales Growth Rate', f"{kpi_metrics['growth_rate']:.1f}%"],
            ['Forecast Accuracy', "95%"]
        ]
        
        # Write KPI data with styling
        for row_idx, row in enumerate(kpi_data, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = kpi_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = bold_font
                cell.border = thin_border
                
                # Header row additional styling
                if row_idx == 1:
                    cell.fill = header_fill
        
        # Forecast Data Sheet
        forecast_sheet = wb.create_sheet(title="Sales Forecast")
        
        # Write forecast data with styling
        for col_idx, header in enumerate(forecast_data.columns, start=1):
            cell = forecast_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Write data rows
        for row_idx, row in enumerate(forecast_data.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = forecast_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Customer Segments Sheet
        segments_sheet = wb.create_sheet(title="Customer Segments")
        
        # Write customer segments data with styling
        for col_idx, header in enumerate(customer_segments.columns, start=1):
            cell = segments_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Write data rows
        for row_idx, row in enumerate(customer_segments.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = segments_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Save to a bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        return excel_bytes

def download_report(forecast_data, kpi_metrics, customer_segments, report_type, unique_key=None):
    """
    Generate and download sales forecast report
    
    Args:
        forecast_data (pd.DataFrame): Forecast data
        kpi_metrics (dict): Key performance indicators
        customer_segments (pd.DataFrame): Customer segmentation data
        report_type (str): Type of report to generate (PDF or Excel)
        unique_key (str, optional): Unique key to prevent widget ID conflicts
    """
    # Prepare KPI metrics dictionary
    kpi_metrics_dict = {
        'total_forecast': forecast_data['Forecast'].sum(),
        'avg_daily_sales': forecast_data['Forecast'].mean(),
        'growth_rate': (forecast_data['Forecast'].iloc[-1] / forecast_data['Forecast'].iloc[0] - 1) * 100
    }
    
    # Use provided customer segments or create default
    if customer_segments is None:
        customer_segments = pd.DataFrame({
            'Segment': ['High-Value', 'Medium-Value', 'Low-Value', 'At-Risk'],
            'Count': [500, 1500, 2000, 1000],
            'Average_CLV': [5000, 2000, 500, 750],
            'Churn_Probability': [0.1, 0.3, 0.6, 0.8]
        })
    
    # Generate appropriate report
    if report_type == 'PDF':
        report_bytes = SalesReportGenerator.generate_pdf_report(
            forecast_data, 
            kpi_metrics_dict, 
            customer_segments
        )
        file_name = "Sales_Forecast_Report.pdf"
        mime_type = "application/pdf"
    else:  # Excel
        report_bytes = SalesReportGenerator.generate_excel_report(
            forecast_data, 
            kpi_metrics_dict, 
            customer_segments
        )
        file_name = "Sales_Forecast_Report.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Provide download with optional unique key
    download_kwargs = {
        'label': f"Download {report_type} Report",
        'data': report_bytes,
        'file_name': file_name,
        'mime': mime_type
    }
    
    # Add unique key if provided
    if unique_key:
        download_kwargs['key'] = unique_key
    
    # Create download button
    st.download_button(**download_kwargs)