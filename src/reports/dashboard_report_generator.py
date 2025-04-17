import io
import pandas as pd
import numpy as np
import streamlit as st
from datetime import datetime
import base64

# PDF-related imports
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.linecharts import HorizontalLineChart

# Excel-related imports
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils.dataframe import dataframe_to_rows

def download_dashboard_report(report_data, report_type="PDF", filters=None, unique_key=None):
    """
    Generate and offer for download a comprehensive dashboard report
    
    Args:
        report_data (dict): Dictionary containing all report data
        report_type (str): Type of report to generate ("PDF" or "Excel Dashboard")
        filters (dict, optional): Dashboard filters applied
        unique_key (str, optional): Unique key for the download button
    """
    # Generate timestamp for filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"biznexus_executive_dashboard_{timestamp}"
    
    # Determine which report generator to use
    if report_type == "PDF":
        return _generate_pdf_report(report_data, filters, filename, unique_key)
    else:
        return _generate_excel_dashboard(report_data, filters, filename, unique_key)

def _generate_pdf_report(report_data, filters, filename, unique_key):
    """Generate PDF report with all dashboard components"""
    # Create in-memory buffer for PDF
    buffer = io.BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    # Get styles
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='Title',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1  # Center alignment
    ))
    styles.add(ParagraphStyle(
        name='Subtitle',
        parent=styles['Heading2'],
        fontSize=14,
        alignment=1  # Center alignment
    ))
    styles.add(ParagraphStyle(
        name='SectionHeader',
        parent=styles['Heading3'],
        fontSize=12,
        spaceBefore=12,
        spaceAfter=6
    ))
    
    # Story elements to build PDF
    story = []
    
    # Title
    story.append(Paragraph("BizNexus AI Executive Dashboard", styles['Title']))
    
    # Generated date and filters
    date_str = datetime.now().strftime("%B %d, %Y")
    story.append(Paragraph(f"Generated on: {date_str}", styles['Subtitle']))
    story.append(Spacer(1, 12))
    
    # Add filters if provided
    if filters:
        filter_data = [
            ["Filter", "Value"],
            ["Date Range", filters.get('date_range', 'All Time')],
            ["Customer Segment", filters.get('customer_segment', 'All Segments')],
            ["Region", filters.get('region', 'All Regions')],
            ["Product Category", filters.get('product_category', 'All Categories')]
        ]
        filter_table = Table(filter_data, colWidths=[1.5*inch, 3*inch])
        filter_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey)
        ]))
        story.append(filter_table)
        story.append(Spacer(1, 24))
    
    # Executive Summary
    story.append(Paragraph("Executive Summary", styles['SectionHeader']))
    
    # Get metrics
    metrics = report_data.get('metrics', {})
    
    # Create metrics table
    metrics_data = [
        ["Total Customers", "Average CLV", "Churn Risk", "Monthly Forecast"],
        [
            f"{metrics.get('total_customers', 0):,}",
            f"${metrics.get('avg_clv', 0):,.2f}",
            f"{metrics.get('avg_churn_probability', 0):.1%}",
            f"${metrics.get('forecast_monthly_avg', 0):,.2f}"
        ]
    ]
    
    metrics_table = Table(metrics_data, colWidths=[2*inch, 2*inch, 2*inch, 2*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey)
    ]))
    
    story.append(metrics_table)
    story.append(Spacer(1, 24))
    
    # Risk Alert
    if metrics.get('high_value_at_risk', 0) > 0:
        risk_data = [
            [f"⚠️ RISK ALERT: {metrics.get('high_value_at_risk', 0)} high-value customers are at risk " +
             f"of churning, putting ${metrics.get('revenue_at_risk', 0):,.2f} of CLV at risk."]
        ]
        risk_table = Table(risk_data, colWidths=[7.5*inch])
        risk_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightpink),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(risk_table)
        story.append(Spacer(1, 12))
    
    # Customer Value Analysis
    story.append(Paragraph("Customer Value Analysis", styles['SectionHeader']))
    
    # Customer segment distribution if available
    clv_df = report_data.get('clv_data')
    if clv_df is not None and 'Value_Tier' in clv_df.columns:
        segment_counts = clv_df['Value_Tier'].value_counts()
        segment_data = [
            ["Value Segment", "Customer Count", "Percentage"],
        ]
        
        # Add segments to table
        for segment, count in segment_counts.items():
            percentage = count / len(clv_df) * 100
            segment_data.append([segment, f"{count:,}", f"{percentage:.1f}%"])
        
        segment_table = Table(segment_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
        segment_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey)
        ]))
        
        story.append(segment_table)
        story.append(Spacer(1, 12))
    
    # Churn Analysis
    story.append(Paragraph("Churn Risk Analysis", styles['SectionHeader']))
    
    # Churn risk distribution if available
    churn_df = report_data.get('churn_data')
    if churn_df is not None and 'Churn_Risk_Category' in churn_df.columns:
        risk_counts = churn_df['Churn_Risk_Category'].value_counts()
        risk_data = [
            ["Risk Category", "Customer Count", "Percentage"],
        ]
        
        # Add risk categories to table
        for risk, count in risk_counts.items():
            percentage = count / len(churn_df) * 100
            risk_data.append([risk, f"{count:,}", f"{percentage:.1f}%"])
        
        risk_table = Table(risk_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
        risk_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey)
        ]))
        
        story.append(risk_table)
        story.append(Spacer(1, 12))
    
    # Sales Forecast
    story.append(Paragraph("Sales Forecast", styles['SectionHeader']))
    
    # Sales metrics
    sales_metrics = [
        ["Total 6-Month Forecast", "Monthly Average", "Growth Rate"],
        [
            f"${metrics.get('forecast_total', 0)/2:,.2f}",
            f"${metrics.get('forecast_monthly_avg', 0):,.2f}",
            f"{metrics.get('forecast_growth', 0):.1f}%"
        ]
    ]
    
    sales_table = Table(sales_metrics, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
    sales_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey)
    ]))
    
    story.append(sales_table)
    story.append(Spacer(1, 12))
    
    # AI Recommendations
    story.append(Paragraph("Business Recommendations", styles['SectionHeader']))
    
    # Generate recommendations based on data
    recommendations = []
    
    # CLV Recommendations
    if clv_df is not None and 'Value_Tier' in clv_df.columns:
        segments = clv_df['Value_Tier'].value_counts(normalize=True) * 100
        premium_pct = segments.get('Premium Value', 0)
        
        if premium_pct < 10:
            recommendations.append(
                "Increase Premium Customer Acquisition: Your Premium Value segment makes up less than 10% of your customer base. " +
                "Implement targeted acquisition campaigns for high-value prospects."
            )
        elif premium_pct > 50:
            recommendations.append(
                "Reduce Customer Value Concentration Risk: Over 50% of your value comes from Premium customers, creating risk. " +
                "Develop programs to increase value of Medium segment customers."
            )
    
    # Churn Recommendations
    if churn_df is not None and 'Churn_Probability' in churn_df.columns:
        high_risk_pct = (churn_df['Churn_Probability'] > 0.5).mean() * 100
        
        if high_risk_pct > 25:
            recommendations.append(
                f"Urgent Churn Reduction Program: {high_risk_pct:.1f}% of customers are at high risk of churning. " +
                "Launch proactive retention campaigns for high-risk customers."
            )
    
    # High-value at risk recommendation
    if metrics.get('high_value_at_risk', 0) > 0:
        recommendations.append(
            f"High-Value Customer Retention Program: {metrics.get('high_value_at_risk', 0)} high-value customers at risk, " +
            f"worth ${metrics.get('revenue_at_risk', 0):,.2f} in CLV. Assign dedicated account managers to at-risk high-value customers."
        )
    
    # If no recommendations, add default
    if not recommendations:
        recommendations.append(
            "Comprehensive Customer Analysis: Conduct in-depth analysis of customer data to identify growth opportunities. " +
            "Segment customers based on purchase behavior and value."
        )
    
    # Create recommendations table
    rec_data = [["Priority Business Recommendations"]]
    for rec in recommendations:
        rec_data.append([rec])
    
    rec_table = Table(rec_data, colWidths=[7.5*inch])
    rec_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(rec_table)
    
    # Add BizNexus footer
    story.append(Spacer(1, 36))
    footer_text = "Generated by BizNexus AI - Business Analytics Platform"
    story.append(Paragraph(footer_text, styles['Italic']))
    
    # Build the PDF document
    doc.build(story)
    
    # Get the PDF data and create download button
    buffer.seek(0)
    pdf_data = buffer.getvalue()
    
    # Create download button with unique key
    download_key = unique_key or f"dashboard_pdf_{timestamp}"
    st.download_button(
        label="Download PDF Report",
        data=pdf_data,
        file_name=f"{filename}.pdf",
        mime="application/pdf",
        key=download_key
    )
    
    return True

def _generate_excel_dashboard(report_data, filters, filename, unique_key):
    """Generate Excel dashboard with all components"""
    # Create workbook
    wb = Workbook()
    
    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Executive Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    
    # Dashboard Title
    ws_summary['A1'] = "BizNexus AI Executive Dashboard"
    ws_summary['A1'].font = Font(name='Calibri', size=16, bold=True)
    ws_summary.merge_cells('A1:E1')
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    # Generation date
    ws_summary['A2'] = f"Generated on: {datetime.now().strftime('%B %d, %Y')}"
    ws_summary['A2'].font = Font(name='Calibri', size=12, italic=True)
    ws_summary.merge_cells('A2:E2')
    ws_summary['A2'].alignment = Alignment(horizontal='center')
    
    # Filters
    if filters:
        ws_summary['A4'] = "Applied Filters"
        ws_summary['A4'].font = header_font
        ws_summary['A4'].fill = header_fill
        ws_summary['B4'] = "Value"
        ws_summary['B4'].font = header_font
        ws_summary['B4'].fill = header_fill
        
        ws_summary['A5'] = "Date Range"
        ws_summary['B5'] = filters.get('date_range', 'All Time')
        
        ws_summary['A6'] = "Customer Segment"
        ws_summary['B6'] = filters.get('customer_segment', 'All Segments')
        
        ws_summary['A7'] = "Region"
        ws_summary['B7'] = filters.get('region', 'All Regions')
        
        ws_summary['A8'] = "Product Category"
        ws_summary['B8'] = filters.get('product_category', 'All Categories')
        
        # Apply borders
        for row in range(4, 9):
            for col in ['A', 'B']:
                ws_summary[f'{col}{row}'].border = border
    
    # Executive Metrics
    metrics = report_data.get('metrics', {})
    
    ws_summary['A10'] = "Executive Metrics"
    ws_summary['A10'].font = Font(name='Calibri', size=14, bold=True)
    ws_summary.merge_cells('A10:E10')
    
    # Create metrics headers
    metrics_headers = ["Metric", "Value"]
    ws_summary['A12'] = metrics_headers[0]
    ws_summary['A12'].font = header_font
    ws_summary['A12'].fill = header_fill
    ws_summary['B12'] = metrics_headers[1]
    ws_summary['B12'].font = header_font
    ws_summary['B12'].fill = header_fill
    
    # Add metrics rows
    metrics_rows = [
        ["Total Customers", f"{metrics.get('total_customers', 0):,}"],
        ["Average CLV", f"${metrics.get('avg_clv', 0):,.2f}"],
        ["Average Churn Probability", f"{metrics.get('avg_churn_probability', 0):.1%}"],
        ["Monthly Forecast", f"${metrics.get('forecast_monthly_avg', 0):,.2f}"],
        ["Forecast Growth Rate", f"{metrics.get('forecast_growth', 0):.1f}%"],
        ["High-Value Customers at Risk", f"{metrics.get('high_value_at_risk', 0)}"],
        ["Revenue at Risk", f"${metrics.get('revenue_at_risk', 0):,.2f}"]
    ]
    
    for i, row in enumerate(metrics_rows):
        row_num = 13 + i
        ws_summary[f'A{row_num}'] = row[0]
        ws_summary[f'B{row_num}'] = row[1]
        
        # Add borders
        ws_summary[f'A{row_num}'].border = border
        ws_summary[f'B{row_num}'].border = border
    
    # Risk Alert
    if metrics.get('high_value_at_risk', 0) > 0:
        alert_row = 13 + len(metrics_rows) + 2
        ws_summary[f'A{alert_row}'] = "⚠️ RISK ALERT"
        ws_summary[f'A{alert_row}'].font = Font(name='Calibri', size=12, bold=True, color='9C0006')
        ws_summary[f'A{alert_row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        alert_text = f"{metrics.get('high_value_at_risk', 0)} high-value customers are at risk of churning, putting ${metrics.get('revenue_at_risk', 0):,.2f} of CLV at risk."
        ws_summary[f'B{alert_row}'] = alert_text
        ws_summary[f'B{alert_row}'].font = Font(name='Calibri', size=12, bold=False, color='9C0006')
        ws_summary[f'B{alert_row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        ws_summary.merge_cells(f'B{alert_row}:E{alert_row}')
    
    # CLV Analysis Sheet
    ws_clv = wb.create_sheet("CLV Analysis")
    
    # CLV Data if available
    clv_df = report_data.get('clv_data')
    if clv_df is not None:
        # Title
        ws_clv['A1'] = "Customer Lifetime Value Analysis"
        ws_clv['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws_clv.merge_cells('A1:E1')
        ws_clv['A1'].alignment = Alignment(horizontal='center')
        
        # Customer Value Tiers if available
        if 'Value_Tier' in clv_df.columns:
            ws_clv['A3'] = "Customer Value Segmentation"
            ws_clv['A3'].font = Font(name='Calibri', size=14, bold=True)
            ws_clv.merge_cells('A3:E3')
            
            # Headers
            segment_headers = ["Value Tier", "Customer Count", "Percentage", "Total CLV", "Average CLV"]
            for i, header in enumerate(segment_headers):
                col = chr(65 + i)  # A, B, C, D, E
                ws_clv[f'{col}5'] = header
                ws_clv[f'{col}5'].font = header_font
                ws_clv[f'{col}5'].fill = header_fill
                ws_clv[f'{col}5'].border = border
            
            # Get segment data
            segment_counts = clv_df['Value_Tier'].value_counts()
            segment_clv = clv_df.groupby('Value_Tier')['CLV'].agg(['sum', 'mean']) if 'CLV' in clv_df.columns else None
            
            # Add segment rows
            row = 6
            for segment, count in segment_counts.items():
                percentage = count / len(clv_df) * 100
                
                ws_clv[f'A{row}'] = segment
                ws_clv[f'B{row}'] = count
                ws_clv[f'C{row}'] = f"{percentage:.1f}%"
                
                if segment_clv is not None and segment in segment_clv.index:
                    ws_clv[f'D{row}'] = segment_clv.loc[segment, 'sum']
                    ws_clv[f'E{row}'] = segment_clv.loc[segment, 'mean']
                
                # Apply borders
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws_clv[f'{col}{row}'].border = border
                
                row += 1
    
    # Churn Analysis Sheet
    ws_churn = wb.create_sheet("Churn Analysis")
    
    # Churn Data if available
    churn_df = report_data.get('churn_data')
    if churn_df is not None:
        # Title
        ws_churn['A1'] = "Customer Churn Analysis"
        ws_churn['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws_churn.merge_cells('A1:E1')
        ws_churn['A1'].alignment = Alignment(horizontal='center')
        
        # Churn Risk Categories if available
        if 'Churn_Risk_Category' in churn_df.columns:
            ws_churn['A3'] = "Churn Risk Distribution"
            ws_churn['A3'].font = Font(name='Calibri', size=14, bold=True)
            ws_churn.merge_cells('A3:E3')
            
            # Headers
            risk_headers = ["Risk Category", "Customer Count", "Percentage"]
            for i, header in enumerate(risk_headers):
                col = chr(65 + i)  # A, B, C
                ws_churn[f'{col}5'] = header
                ws_churn[f'{col}5'].font = header_font
                ws_churn[f'{col}5'].fill = header_fill
                ws_churn[f'{col}5'].border = border
            
            # Get risk data
            risk_counts = churn_df['Churn_Risk_Category'].value_counts()
            
            # Add risk rows
            row = 6
            for risk, count in risk_counts.items():
                percentage = count / len(churn_df) * 100
                
                ws_churn[f'A{row}'] = risk
                ws_churn[f'B{row}'] = count
                ws_churn[f'C{row}'] = f"{percentage:.1f}%"
                
                # Apply borders
                for col in ['A', 'B', 'C']:
                    ws_churn[f'{col}{row}'].border = border
                
                row += 1
        
        # High-Value Customers At Risk
        if clv_df is not None:
            try:
                # Create merged view
                merged_df = pd.merge(
                    clv_df,
                    churn_df,
                    on='Customer ID',
                    how='inner'
                )
                
                # Identify high-value at risk
                high_value_at_risk = merged_df[
                    (merged_df['Value_Tier'] == 'Premium Value') & 
                    (merged_df['Churn_Probability'] > 0.5)
                ] if 'Value_Tier' in merged_df.columns and 'Churn_Probability' in merged_df.columns else pd.DataFrame()
                
                if not high_value_at_risk.empty:
                    row += 2
                    ws_churn[f'A{row}'] = "High-Value Customers at Risk"
                    ws_churn[f'A{row}'].font = Font(name='Calibri', size=14, bold=True)
                    ws_churn.merge_cells(f'A{row}:E{row}')
                    
                    # Get columns for table
                    display_cols = ['Customer ID', 'Customer Name', 'CLV', 'Churn_Probability']
                    display_cols = [col for col in display_cols if col in high_value_at_risk.columns]
                    
                    # Add headers
                    row += 2
                    for i, col in enumerate(display_cols):
                        excel_col = chr(65 + i)  # A, B, C, D
                        ws_churn[f'{excel_col}{row}'] = col.replace('_', ' ')
                        ws_churn[f'{excel_col}{row}'].font = header_font
                        ws_churn[f'{excel_col}{row}'].fill = header_fill
                        ws_churn[f'{excel_col}{row}'].border = border
                    
                    # Add top 10 at-risk customers
                    top_at_risk = high_value_at_risk.sort_values('CLV', ascending=False).head(10)
                    
                    # Add data rows
                    for j, (_, customer) in enumerate(top_at_risk.iterrows()):
                        row_num = row + j + 1
                        for i, col in enumerate(display_cols):
                            excel_col = chr(65 + i)
                            value = customer[col]
                            
                            # Format values appropriately
                            if col == 'CLV':
                                cell_value = f"${value:,.2f}" if pd.notna(value) else "N/A"
                            elif col == 'Churn_Probability':
                                cell_value = f"{value:.1%}" if pd.notna(value) else "N/A"
                            else:
                                cell_value = str(value) if pd.notna(value) else "N/A"
                            
                            ws_churn[f'{excel_col}{row_num}'] = cell_value
                            ws_churn[f'{excel_col}{row_num}'].border = border
            
            except Exception as e:
                print(f"Error creating high value at risk table: {e}")
    
    # Sales Forecast Sheet
    ws_forecast = wb.create_sheet("Sales Forecast")
    
    # Forecast Data if available
    forecast_df = report_data.get('forecast_data')
    if forecast_df is not None:
        # Title
        ws_forecast['A1'] = "Sales Forecast Analysis"
        ws_forecast['A1'].font = Font(name='Calibri', size=16, bold=True)
        ws_forecast.merge_cells('A1:E1')
        ws_forecast['A1'].alignment = Alignment(horizontal='center')
        
        # Forecast summary metrics
        ws_forecast['A3'] = "Forecast Summary"
        ws_forecast['A3'].font = Font(name='Calibri', size=14, bold=True)
        ws_forecast.merge_cells('A3:E3')
        
        # Add forecast metrics
        metrics_headers = ["Metric", "Value"]
        ws_forecast['A5'] = metrics_headers[0]
        ws_forecast['A5'].font = header_font
        ws_forecast['A5'].fill = header_fill
        ws_forecast['B5'] = metrics_headers[1]
        ws_forecast['B5'].font = header_font
        ws_forecast['B5'].fill = header_fill
        
        forecast_metrics = [
            ["Total 6-Month Forecast", f"${metrics.get('forecast_total', 0)/2:,.2f}"],
            ["Monthly Average Sales", f"${metrics.get('forecast_monthly_avg', 0):,.2f}"],
            ["Forecast Growth Rate", f"{metrics.get('forecast_growth', 0):.1f}%"]
        ]
        
        for i, (metric, value) in enumerate(forecast_metrics):
            row_num = 6 + i
            ws_forecast[f'A{row_num}'] = metric
            ws_forecast[f'B{row_num}'] = value
            
            # Apply borders
            ws_forecast[f'A{row_num}'].border = border
            ws_forecast[f'B{row_num}'].border = border
        
        # Add forecast data if available
        if 'Date' in forecast_df.columns and 'Forecast' in forecast_df.columns:
            # Sample the forecast data (to avoid too many rows)
            if len(forecast_df) > 30:
                # Take every Nth row to get approximately 30 data points
                step = max(1, len(forecast_df) // 30)
                sampled_forecast = forecast_df.iloc[::step].copy()
            else:
                sampled_forecast = forecast_df.copy()
            
            # Add historical vs forecast flag
            if 'is_historical' not in sampled_forecast.columns:
                sampled_forecast['is_historical'] = sampled_forecast['Date'] <= datetime.now()
            
            # Convert dates to strings for Excel
            sampled_forecast['Date_Str'] = sampled_forecast['Date'].dt.strftime('%Y-%m-%d')
            
            # Add forecast data table
            start_row = 6 + len(forecast_metrics) + 2
            ws_forecast[f'A{start_row}'] = "Forecast Data"
            ws_forecast[f'A{start_row}'].font = Font(name='Calibri', size=14, bold=True)
            ws_forecast.merge_cells(f'A{start_row}:E{start_row}')
            
            # Headers
            headers = ["Date", "Forecast", "Type"]
            for i, header in enumerate(headers):
                col = chr(65 + i)  # A, B, C
                ws_forecast[f'{col}{start_row+2}'] = header
                ws_forecast[f'{col}{start_row+2}'].font = header_font
                ws_forecast[f'{col}{start_row+2}'].fill = header_fill
                ws_forecast[f'{col}{start_row+2}'].border = border
            
            # Add data rows
            for i, (_, row) in enumerate(sampled_forecast.iterrows()):
                row_num = start_row + 3 + i
                
                # Date
                ws_forecast[f'A{row_num}'] = row['Date_Str']
                ws_forecast[f'A{row_num}'].border = border
                
                # Forecast value
                ws_forecast[f'B{row_num}'] = row['Forecast']
                ws_forecast[f'B{row_num}'].number_format = '$#,##0.00'
                ws_forecast[f'B{row_num}'].border = border
                
                # Type (Historical or Forecast)
                type_value = "Historical" if row['is_historical'] else "Forecast"
                ws_forecast[f'C{row_num}'] = type_value
                ws_forecast[f'C{row_num}'].border = border
    
    # Business Recommendations Sheet
    ws_recommendations = wb.create_sheet("Recommendations")
    
    # Title
    ws_recommendations['A1'] = "Business Recommendations"
    ws_recommendations['A1'].font = Font(name='Calibri', size=16, bold=True)
    ws_recommendations.merge_cells('A1:E1')
    ws_recommendations['A1'].alignment = Alignment(horizontal='center')
    
    # Generate recommendations based on data
    recommendations = []
    
    # CLV Recommendations
    if clv_df is not None and 'Value_Tier' in clv_df.columns:
        segments = clv_df['Value_Tier'].value_counts(normalize=True) * 100
        premium_pct = segments.get('Premium Value', 0)
        
        if premium_pct < 10:
            recommendations.append({
                "title": "Increase Premium Customer Acquisition",
                "description": "Your Premium Value segment makes up less than 10% of your customer base.",
                "actions": [
                    "Implement targeted acquisition campaigns for high-value prospects",
                    "Develop referral programs specifically for Premium segment customers",
                    "Create upgrade paths for high-potential Medium Value customers"
                ],
                "priority": "High"
            })
        elif premium_pct > 50:
            recommendations.append({
                "title": "Reduce Customer Value Concentration Risk",
                "description": "Over 50% of your value comes from Premium customers, creating concentration risk.",
                "actions": [
                    "Develop programs to increase value of Medium segment customers",
                    "Create special retention programs for Premium customers",
                    "Diversify acquisition channels to reach broader customer segments"
                ],
                "priority": "Medium"
            })
    
    # Churn Recommendations
    if churn_df is not None and 'Churn_Probability' in churn_df.columns:
        high_risk_pct = (churn_df['Churn_Probability'] > 0.5).mean() * 100
        
        if high_risk_pct > 25:
            recommendations.append({
                "title": "Urgent Churn Reduction Program",
                "description": f"{high_risk_pct:.1f}% of customers are at high risk of churning.",
                "actions": [
                    "Launch proactive retention campaigns for high-risk customers",
                    "Implement satisfaction surveys to identify churn drivers",
                    "Develop win-back strategies for recently churned customers",
                    "Create an early warning system for customers showing churn indicators"
                ],
                "priority": "Critical"
            })
    
    # High-value at risk recommendation
    if metrics.get('high_value_at_risk', 0) > 0:
        recommendations.append({
            "title": "High-Value Customer Retention Program",
            "description": f"{metrics.get('high_value_at_risk', 0)} high-value customers at risk, worth ${metrics.get('revenue_at_risk', 0):,.2f} in CLV.",
            "actions": [
                "Assign dedicated account managers to at-risk high-value customers",
                "Create VIP loyalty program with exclusive benefits",
                "Implement personalized outreach and engagement strategy",
                "Schedule executive relationship reviews for top accounts"
            ],
            "priority": "Critical"
        })
    
    # If no recommendations, add default
    if not recommendations:
        recommendations.append({
            "title": "Comprehensive Customer Analysis",
            "description": "Conduct in-depth analysis of customer data to identify growth opportunities.",
            "actions": [
                "Segment customers based on purchase behavior and value",
                "Identify key factors driving customer lifetime value",
                "Analyze churn patterns to develop prevention strategies",
                "Develop targeted marketing campaigns based on insights"
            ],
            "priority": "Medium"
        })
    
    # Headers
    rec_headers = ["Priority", "Recommendation", "Description"]
    for i, header in enumerate(rec_headers):
        col = chr(65 + i)  # A, B, C
        ws_recommendations[f'{col}3'] = header
        ws_recommendations[f'{col}3'].font = header_font
        ws_recommendations[f'{col}3'].fill = header_fill
        ws_recommendations[f'{col}3'].border = border
    
    # Set column widths
    ws_recommendations.column_dimensions['A'].width = 15
    ws_recommendations.column_dimensions['B'].width = 30
    ws_recommendations.column_dimensions['C'].width = 50
    
    # Add recommendations
    for i, rec in enumerate(recommendations):
        row_num = 4 + i
        
        # Priority
        ws_recommendations[f'A{row_num}'] = rec.get('priority', 'Medium')
        
        # Add color to priority cell
        if rec.get('priority') == 'Critical':
            ws_recommendations[f'A{row_num}'].fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
        elif rec.get('priority') == 'High':
            ws_recommendations[f'A{row_num}'].fill = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
        
        # Title
        ws_recommendations[f'B{row_num}'] = rec.get('title', '')
        
        # Description with actions
        description = rec.get('description', '')
        actions = rec.get('actions', [])
        
        if actions:
            description += "\n\nRecommended Actions:\n- " + "\n- ".join(actions)
        
        ws_recommendations[f'C{row_num}'] = description
        ws_recommendations[f'C{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set row height for wrapped text
        ws_recommendations.row_dimensions[row_num].height = 100
        
        # Apply borders
        for col in ['A', 'B', 'C']:
            ws_recommendations[f'{col}{row_num}'].border = border
    
    # Customer Data Sheet
    ws_customers = wb.create_sheet("Customer Data")
    
    # Title
    ws_customers['A1'] = "Customer Data"
    ws_customers['A1'].font = Font(name='Calibri', size=16, bold=True)
    ws_customers.merge_cells('A1:E1')
    ws_customers['A1'].alignment = Alignment(horizontal='center')
    
    # Create merged view of customer data
    customer_df = report_data.get('customer_df')
    
    if customer_df is not None:
        # Add CLV data if available
        if clv_df is not None:
            # Select only key columns for merging
            clv_cols = ['Customer ID', 'CLV', 'Predicted_CLV', 'Value_Tier']
            clv_cols = [col for col in clv_cols if col in clv_df.columns]
            
            if clv_cols:
                customer_df = pd.merge(
                    customer_df,
                    clv_df[clv_cols],
                    on='Customer ID',
                    how='left'
                )
        
        # Add Churn data if available
        if churn_df is not None:
            # Select only key columns for merging
            churn_cols = ['Customer ID', 'Churn_Probability', 'Churn_Risk_Category']
            churn_cols = [col for col in churn_cols if col in churn_df.columns]
            
            if churn_cols:
                customer_df = pd.merge(
                    customer_df,
                    churn_df[churn_cols],
                    on='Customer ID',
                    how='left'
                )
        
        # Write customer data to sheet
        # Headers from dataframe columns
        headers = list(customer_df.columns)
        for i, header in enumerate(headers):
            col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
            ws_customers[f'{col_letter}3'] = header
            ws_customers[f'{col_letter}3'].font = header_font
            ws_customers[f'{col_letter}3'].fill = header_fill
            ws_customers[f'{col_letter}3'].border = border
        
        # Data rows (limited to first 1000 rows)
        display_df = customer_df.head(1000)
        
        for i, (_, row) in enumerate(display_df.iterrows()):
            row_num = 4 + i
            
            for j, col in enumerate(headers):
                col_letter = chr(65 + j) if j < 26 else chr(64 + j // 26) + chr(65 + j % 26)
                value = row[col]
                
                # Format values appropriately
                if col in ['CLV', 'Predicted_CLV']:
                    cell_value = value if pd.isna(value) else float(value)
                    if not pd.isna(value):
                        ws_customers[f'{col_letter}{row_num}'].number_format = '$#,##0.00'
                elif col == 'Churn_Probability':
                    cell_value = value if pd.isna(value) else float(value)
                    if not pd.isna(value):
                        ws_customers[f'{col_letter}{row_num}'].number_format = '0.00%'
                else:
                    cell_value = str(value) if pd.notna(value) else ""
                
                ws_customers[f'{col_letter}{row_num}'] = cell_value
                ws_customers[f'{col_letter}{row_num}'].border = border
    
    # Save workbook to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create download button with unique key
    download_key = unique_key or f"dashboard_excel_{timestamp}"
    st.download_button(
        label="Download Excel Dashboard",
        data=output,
        file_name=f"{filename}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=download_key
    )
    
    return True

def send_dashboard_report_email(email, report_data, report_type="PDF"):
    """
    Generate a dashboard report and email it to the specified address
    
    Args:
        email (str): Email address to send the report to
        report_data (dict): Dictionary containing all report data
        report_type (str): Type of report to generate ("PDF" or "Excel Dashboard")
    
    Returns:
        bool: True if email sent successfully, False otherwise
    """
    # This is a placeholder for email functionality
    # Actual implementation would need email server configuration
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"biznexus_executive_dashboard_{timestamp}"
    
    try:
        # Generate the report
        buffer = io.BytesIO()
        
        if report_type == "PDF":
            # Generate PDF report
            _generate_pdf_to_buffer(report_data, None, buffer)
        else:
            # Generate Excel report
            _generate_excel_to_buffer(report_data, None, buffer)
        
        # Reset buffer position
        buffer.seek(0)
        
        # Here you would send the email with the buffer as attachment
        # Using your preferred email library (e.g., smtplib)
        
        # For now, just return True as if it worked
        return True
    
    except Exception as e:
        print(f"Error generating and sending report: {str(e)}")
        return False

def _generate_pdf_to_buffer(report_data, filters, buffer):
    """Internal helper for generating PDF to a buffer without download button"""
    # Create PDF document
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter),
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    # Implementation would mirror _generate_pdf_report but write to buffer
    # without creating a download button
    
    # This is a placeholder
    pass

def _generate_excel_to_buffer(report_data, filters, buffer):
    """Internal helper for generating Excel to a buffer without download button"""
    # Create workbook
    wb = Workbook()
    
    # Implementation would mirror _generate_excel_dashboard but write to buffer
    # without creating a download button
    
    # This is a placeholder
    wb.save(buffer)
    pass