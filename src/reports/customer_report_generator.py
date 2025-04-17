import io
import pandas as pd
import streamlit as st

# PDF Generation
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# Excel Generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

class CustomerReportGenerator:
    @staticmethod
    def generate_pdf_report(customer_data, customer_metrics, customer_segments):
        """
        Generate a comprehensive customer PDF report
        
        Args:
            customer_data (pd.DataFrame): Customer detailed data
            customer_metrics (dict): Key customer performance indicators
            customer_segments (pd.DataFrame): Customer segmentation data
        
        Returns:
            bytes: PDF report content
        """
        # Ensure all required keys exist in customer_metrics
        required_keys = ['customer_lifetime_value', 'churn_probability', 'retention_rate']
        for key in required_keys:
            if key not in customer_metrics:
                customer_metrics[key] = 0  # Default value if missing

        # Create a file-like buffer to receive PDF data
        buffer = io.BytesIO()
        
        # Create PDF document
        pdf = SimpleDocTemplate(buffer, pagesize=letter)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        heading_style = styles['Heading2']
        normal_style = styles['Normal']
        
        # Build PDF content
        story = []
        
        # Title
        report_title = customer_metrics.get('custom_title', 'Customer Report')
        story.append(Paragraph(report_title, title_style))
        story.append(Spacer(1, 12))
        
        # Customer KPI Section
        story.append(Paragraph("Customer Performance Indicators", heading_style))
        
        # Prepare KPI table data
        kpi_data = [
            ['Metric', 'Value'],
            ['Customer Lifetime Value', f"${customer_metrics['customer_lifetime_value']:,.0f}"],
            ['Churn Probability', f"{customer_metrics['churn_probability']:.1%}"],
            ['Retention Rate', f"{customer_metrics['retention_rate']:.1%}"],
            ['Customer Satisfaction', "4.5/5"]
        ]
        
        # Create KPI table
        kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        
        story.append(kpi_table)
        story.append(Spacer(1, 12))
        
        # Customer Data Section
        story.append(Paragraph("Customer Detailed Information", heading_style))
        
        # Prepare customer data table
        customer_data_list = [customer_data.columns.tolist()] + customer_data.head(10).values.tolist()
        customer_table = Table(customer_data_list, colWidths=[1.5*inch]*len(customer_data.columns))
        customer_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        
        story.append(customer_table)
        story.append(Spacer(1, 12))
        
        # Customer Segments Section
        story.append(Paragraph("Customer Segmentation", heading_style))
        
        # Prepare customer segments table
        segment_data = [customer_segments.columns.tolist()] + customer_segments.values.tolist()
        segment_table = Table(segment_data, colWidths=[1.5*inch]*len(customer_segments.columns))
        segment_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        
        story.append(segment_table)
        
        # Generate PDF
        pdf.build(story)
        
        # Get the value of the BytesIO buffer and write it to the response
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes

    @staticmethod
    def generate_excel_report(customer_data, customer_metrics, customer_segments):
        """
        Generate a comprehensive customer Excel report
        
        Args:
            customer_data (pd.DataFrame): Customer detailed data
            customer_metrics (dict): Key customer performance indicators
            customer_segments (pd.DataFrame): Customer segmentation data
        
        Returns:
            bytes: Excel report content
        """
        # Ensure all required keys exist in customer_metrics
        required_keys = ['customer_lifetime_value', 'churn_probability', 'retention_rate']
        for key in required_keys:
            if key not in customer_metrics:
                customer_metrics[key] = 0  # Default value if missing

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        
        # Customer Metrics Sheet
        metrics_sheet = wb.active
        metrics_sheet.title = "Customer Metrics"
        
        # Headers and styling
        bold_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Customer Metrics Data
        metrics_data = [
            ['Metric', 'Value'],
            ['Customer Lifetime Value', f"${customer_metrics['customer_lifetime_value']:,.0f}"],
            ['Churn Probability', f"{customer_metrics['churn_probability']:.1%}"],
            ['Retention Rate', f"{customer_metrics['retention_rate']:.1%}"],
            ['Customer Satisfaction', "4.5/5"]
        ]
        
        # Write customer metrics data with styling
        for row_idx, row in enumerate(metrics_data, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = metrics_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = bold_font
                cell.border = thin_border
                
                # Header row additional styling
                if row_idx == 1:
                    cell.fill = header_fill
        
        # Customer Data Sheet
        customer_sheet = wb.create_sheet(title="Customer Details")
        
        # Write customer data with styling
        for col_idx, header in enumerate(customer_data.columns, start=1):
            cell = customer_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Write data rows
        for row_idx, row in enumerate(customer_data.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = customer_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Customer Segments Sheet
        segments_sheet = wb.create_sheet(title="Customer Segments")
        
        # Write customer segments data with styling
        for col_idx, header in enumerate(customer_segments.columns, start=1):
            cell = segments_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Write data rows
        for row_idx, row in enumerate(customer_segments.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = segments_sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
        
        # Save to a bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        return excel_bytes

def download_customer_report(forecast_data, kpi_metrics, customer_segments, report_type, unique_key=None):
    """
    Generate and download customer report
    
    Args:
        forecast_data (pd.DataFrame): DataFrame containing customer data
        kpi_metrics (dict): Key performance indicators
        customer_segments (pd.DataFrame): Customer segmentation data
        report_type (str): Type of report to generate (PDF or Excel)
        unique_key (str, optional): Unique key to prevent widget ID conflicts
    """
    # Prepare customer metrics dictionary
    customer_metrics_dict = {
        'customer_lifetime_value': kpi_metrics.get('customer_value', 0),
        'churn_probability': kpi_metrics.get('churn_risk', 0),
        'retention_rate': kpi_metrics.get('retention_score', 1 - kpi_metrics.get('churn_risk', 0)),
        'custom_title': kpi_metrics.get('custom_title', 'Customer Report')
    }
    
    # Use provided customer segments or create default
    if customer_segments is None:
        customer_segments = pd.DataFrame({
            'Segment': ['High-Value', 'Medium-Value', 'Low-Value', 'At-Risk'],
            'Count': [500, 1500, 2000, 1000],
            'Average_CLV': [5000, 2000, 500, 750],
            'Churn_Probability': [0.1, 0.3, 0.6, 0.8]
        })
    
    # Generate appropriate report
    if report_type == 'PDF':
        report_bytes = CustomerReportGenerator.generate_pdf_report(
            forecast_data, 
            customer_metrics_dict, 
            customer_segments
        )
        file_name = f"{customer_metrics_dict['custom_title']}.pdf"
        mime_type = "application/pdf"
    else:  # Excel
        report_bytes = CustomerReportGenerator.generate_excel_report(
            forecast_data, 
            customer_metrics_dict, 
            customer_segments
        )
        file_name = f"{customer_metrics_dict['custom_title']}.xlsx"
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Provide download with optional unique key
    download_kwargs = {
        'label': f"Download {report_type} Report",
        'data': report_bytes,
        'file_name': file_name,
        'mime': mime_type
    }
    
    # Add unique key if provided
    if unique_key:
        download_kwargs['key'] = unique_key
    
    # Create download button
    st.download_button(**download_kwargs)